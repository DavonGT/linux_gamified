from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Task, Chapter, Mission
from openpyxl import load_workbook

admin.site.register(Chapter)
admin.site.register(Mission)

@admin.register(Task)
class TaskAdmin(admin.ModelAdmin):
    list_display = ('task', 'difficulty', 'category')  # Display task and difficulty
    change_list_template = "admin/game/task_changelist.html"

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('upload-excel/', self.upload_excel, name='upload_excel'),
        ]
        return custom_urls + urls

    def upload_excel(self, request):
        if request.method == "POST":
            excel_file = request.FILES["excel_file"]
            try:
                # Load the Excel file
                workbook = load_workbook(excel_file)
                sheet = workbook.active

                # Ensure the required columns exist
                headers = [cell.value for cell in sheet[1]]
                required_columns = {'task', 'correct_commands', 'difficulty', 'chapter_number'}
                if not required_columns.issubset(headers):
                    messages.error(request, "Excel file must contain 'task', 'correct_commands', 'difficulty', and 'chapter_number' columns.")
                    return redirect("..")

                # Map column headers to indices
                header_indices = {header: idx for idx, header in enumerate(headers)}

                # Print header indices for debugging
                print("Header Indices:", header_indices)

                # Iterate through rows and create Task, Mission, and Chapter objects
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    # Print each row for debugging
                    print("Row data:", row)

                    # Ensure chapter_number exists and is valid
                    chapter_number = row[header_indices.get('chapter_number', -1)]
                    
                    # Print chapter_number for debugging
                    print("Chapter number:", chapter_number)

                    # Check if chapter_number is missing or invalid
                    if not chapter_number:
                        print("Skipping row due to missing chapter_number.")
                        continue  

                    chapter_number = str(chapter_number).strip()

                    # Skip rows where chapter_number is not valid
                    if not chapter_number.isdigit():
                        print(f"Skipping row due to invalid chapter_number: {chapter_number}")
                        continue  

                    chapter_number = int(chapter_number)
                    chapter_name = f"Chapter {chapter_number}"

                    chapter, chapter_created = Chapter.objects.get_or_create(
                        number=chapter_number, 
                        defaults={'name': chapter_name}
                    )

                    task, task_created = Task.objects.get_or_create(
                        task=row[header_indices['task']],  # Ensure no duplicate task based on task name
                        defaults={
                            'correct_commands': row[header_indices['correct_commands']],
                            'difficulty': row[header_indices['difficulty']],
                            'category': 'other' if 'category' not in header_indices else row[header_indices['category']],
                        }
                    )

                    mission, mission_created = Mission.objects.get_or_create(
                        chapter=chapter,
                        task=task,  # Link the task to the mission
                        defaults={'chapter': chapter, 'task': task}
                    )

                    # Ensure the mission task is linked (if not linked automatically)
                    if mission_created:
                        mission.task = task
                        mission.save()

                messages.success(request, "Tasks, Chapters, and Missions have been successfully uploaded!")
                return redirect("admin:game_task_changelist")  # Explicit redirect to the task changelist page
            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
                return redirect("..")

        return render(request, "admin/game/upload_excel.html")
